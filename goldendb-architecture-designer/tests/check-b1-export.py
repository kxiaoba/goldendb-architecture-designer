"""Read-only checks of the workbook actually downloaded by the browser test."""
from pathlib import Path
import json
import zipfile
import openpyxl

root = Path(__file__).resolve().parents[2]
directory = root / 'outputs/goldendb-remediation-20260906/evidence'
expected = json.loads((directory / 'b1-export-sheets.json').read_text())
source = max(directory.glob('b1-*.xlsx'), key=lambda p: p.stat().st_mtime)
workbook = openpyxl.load_workbook(source, data_only=False)
differences = []
errors = []
checked = 0
for sheet in expected:
    ws = workbook[sheet['name']]
    for row_number, row in enumerate(sheet['rows'], 1):
        for column, value in enumerate(row, 1):
            actual = ws.cell(row_number, column).value
            checked += 1
            if (actual if actual is not None else '') != (value if value is not None else ''):
                differences.append([sheet['name'], row_number, column, value, actual])
    errors.extend([ws.title, cell.coordinate, cell.value] for row in ws for cell in row if cell.data_type == 'e')
with zipfile.ZipFile(source) as archive:
    zip_error = archive.testzip()
result = {
    'file': source.name, 'sheets': workbook.sheetnames, 'checkedCells': checked,
    'differences': differences, 'formulaErrors': errors, 'zipError': zip_error,
    'pass': not differences and not errors and zip_error is None,
    'scope': '实际文件与当前导出模型逐格一致；本批未改变工作簿布局。不是实库性能或生产架构认证。'
}
(directory / 'b1-export-validation.json').write_text(json.dumps(result, ensure_ascii=False, indent=2))
print(json.dumps(result, ensure_ascii=False, indent=2))
raise SystemExit(0 if result['pass'] else 1)
